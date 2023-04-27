import time
import openpyxl
from tqdm import tqdm
from tabulate import tabulate
import colorama
from colorama import Fore, Style


def printColor(color, string):
    fColor = ''
    if color == 'cyan':
        fColor = Fore.CYAN
    elif color == 'red':
        fColor = Fore.RED
    elif color == 'green':
        fColor = Fore.GREEN
    elif color == 'yellow':
        fColor = Fore.YELLOW
    elif color == 'purple':
        fColor = Fore.MAGENTA 
        
    print(fColor + string + Style.RESET_ALL)


def main():
    colorama.init()
    path_of_excel_file = '../PACKING  - 2023-U9.xlsx'
    
    tab = select_a_tab(path_of_excel_file)
    
    if 'IN' in tab:
        check_in(path_of_excel_file, tab)
    elif 'OUT' in tab:
        check_out(path_of_excel_file, tab)
    
    print("Press any key to continue...")
    input()


def select_a_tab(path_of_excel_file):
    wb = openpyxl.load_workbook(path_of_excel_file)
    worksheet_names = wb.sheetnames
    
    # index = 1
    # for name in worksheet_names:
    #     if 'IN' in name and 'TOTAL' not in name:
    #         printColor('yellow', f'{index} - {name}')
    #         index += 1
    #     if 'OUT' in name and 'TOTAL' not in name:
    #         printColor('green', f'{index} - {name}')
    #         index += 1
    
    ws_list = []
    
    for name in worksheet_names:
        if ('IN' in name or 'OUT' in name) and 'TOTAL' not in name:
            ws_list.append(name)
    
    for i, ws in enumerate(ws_list):
        color = ''
        if 'IN' in ws:
            color = 'yellow'
        elif 'OUT' in ws:
            color = 'green'
        printColor(color, f'{i} - {ws}')
    
    selected_index = -1
    while selected_index < 0 or selected_index > len(ws):
        selected_index = get_input_and_validate()
    
    return ws_list[selected_index]


def get_input_and_validate():
    try:
        selected_index = int(input('Select a worksheet: '))
    except:
        selected_index = 0
    
    return selected_index

def check_in(path_of_excel_file, tab):
    path_of_u9_file = './u9_excel/MO.xlsx'
    items = {}
    items = get_excel_in_records(path_of_excel_file, tab, items)
    items = get_u9_in_records(path_of_u9_file, items)
    # print(items)
    show_result(items)
            

def get_excel_in_records(path_of_excel_file, tab, items):
    # Open the workbook and worksheet
    wb = openpyxl.load_workbook(path_of_excel_file, data_only=True)
    ws = wb[tab]

    # Loop through each row starting from row 3
    print('Reading Excel...')
    for row in tqdm(range(3, ws.max_row + 1)):
        item_id = ws.cell(row=row, column=4).value
        
        if str(item_id) == '0':
            break
        
        # Declare an empty dictionary for each row
        item = {}

        # Get the values of the 'id' and 'excel_qty' columns
        # item['id'] = item_id
        item['excel_qty'] = ws.cell(row=row, column=38).value
        
        item['excel_qty'] = round(item['excel_qty'], 4)
        
        items[item_id] = item

    # Return the list of items
    return items


def get_u9_in_records(path_of_u9_file, items):
    # print(items)
    wb = openpyxl.load_workbook(filename=path_of_u9_file, read_only=True)
    ws = wb[wb.sheetnames[0]]

    id_col = None
    qty_col = None

    for row in ws.iter_rows(min_row=1, max_row=3, min_col=1, max_col=20):
        for cell in row:
            if cell.value == '存货代码' or cell.value == '料品.参考料号2':
                id_col = cell.column
                header_row = cell.row
            elif cell.value == '入库数量(生产单位)' or cell.value == '生产数量':
                qty_col = cell.column
            elif cell.value == None:
                break
    
    print('Reading U9...')
                
    current_row = header_row
    for row in tqdm(ws.iter_rows(min_row=header_row+1, max_row=1000)):
        current_row += 1
        item_id = ws.cell(current_row, column=id_col).value
        u9_qty = ws.cell(current_row, column=qty_col).value
        
        # print(item_id)
        # print(u9_qty)
        
        if item_id == None:
            break

        if item_id in items and 'u9_qty' in items[item_id]:
            items[item_id]['u9_qty'] += u9_qty
        else:
            items[item_id]['u9_qty'] = u9_qty

    return items


def check_out(path_of_excel_file, tab):
    path_of_u9_file = './u9_excel/004.xlsx'
    items = {}
    items = get_excel_out_records(path_of_excel_file, tab, items)
    items = get_u9_out_records(path_of_u9_file, items)
    # print(items)
    show_result(items)


def get_excel_out_records(path_of_excel_file, tab, items):
    # Open the workbook and worksheet
    wb = openpyxl.load_workbook(path_of_excel_file, data_only=True)
    ws = wb[tab]

    # Loop through each row starting from row 3
    print('Reading Excel...')
    for row in tqdm(range(3, ws.max_row + 1)):
        item_id = ws.cell(row=row, column=4).value
        
        if str(item_id) == '0':
            break
        
        # Declare an empty dictionary for each row
        item = {}

        # Get the values of the 'id' and 'excel_qty' columns
        # item['id'] = item_id
        item['excel_qty'] = ws.cell(row=row, column=39).value
        
        item['excel_qty'] = round(item['excel_qty'], 4)
        
        items[item_id] = item

    # Return the list of items
    return items


def get_u9_out_records(path_of_u9_file, items):
    # print(items)
    wb = openpyxl.load_workbook(filename=path_of_u9_file, read_only=True)
    ws = wb[wb.sheetnames[0]]

    id_col = None
    qty_col = None

    for row in ws.iter_rows(min_row=1, max_row=3, min_col=1, max_col=20):
        for cell in row:
            if cell.value == '参考料号2' or cell.value == '代码':
                id_col = cell.column
                header_row = cell.row
            elif cell.value == '现存量(库存单位)':
                qty_col = cell.column

    current_row = header_row
    empty_cell_count = 0
    last_row_enpty = False
    print('Reading U9...')
    for row in tqdm(ws.iter_rows(min_row=header_row+1, max_row=1000)):
        current_row += 1
        # try:
        item_id = ws.cell(row=current_row, column=id_col).value
        u9_qty = ws.cell(row=current_row, column=qty_col).value
        # except:
        #     continue
        
        # print(item_id)
        # print(u9_qty)
        
        if item_id == None or item_id == '':
            if last_row_enpty: empty_cell_count += 1
            last_row_enpty = True
            if empty_cell_count >= 50:
                break
            continue
        
        last_row_enpty = False
        
        if item_id not in items:
            items[item_id] = {}

        if 'u9_qty' in items[item_id]:
            items[item_id]['u9_qty'] += u9_qty
        else:
            items[item_id]['u9_qty'] = u9_qty

    return items


def show_result(items):
    unmatch_items = {}
    
    for key, value in items.items():
        e = value.get("excel_qty", 0)
        u = value.get("u9_qty", 0)
        
        if e == u:
            color = Fore.CYAN
        else:
            color = Fore.RED
            unmatch_items[key] = value
            
        if not (e == 0 and u == 0):
            print(key + color + f' -> Excel: {e} U9: {u}' + Style.RESET_ALL)
            time.sleep(0.15)
    
    if len(unmatch_items) == 0: return
    
    # print('\nFound the following unmatched numbers:')
    # for key, value in unmatch_items.items():
    #     e = value.get("excel_qty", 0)
    #     u = value.get("u9_qty", 0)
    #     print(key + Fore.RED + f' -> Excel: {e} U9: {u}' + Style.RESET_ALL)
    
    output_table = []
    for key, value in unmatch_items.items():
        e = value.get("excel_qty", 0)
        u = value.get("u9_qty", 0)
        # print(key + Fore.RED + f' -> Excel: {e} U9: {u}' + Style.RESET_ALL)
        output_table.append([key, f'Excel: {e}', f'U9: {u}', f'Diff: {round(abs(e - u),3)}'])
        
    print('\nFound the following unmatched numbers:')
    print(tabulate(output_table))


if __name__ == '__main__':
    main()