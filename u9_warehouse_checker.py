import re
import time
import copy
import openpyxl
from tqdm import tqdm
from tabulate import tabulate
import colorama
from colorama import Fore, Style


# path_of_excel = '../INVENTORY 2023.MAY.xlsx'


def printColor(color, string):
    fColor = ''
    if color == 'cyan':
        fColor = Fore.CYAN
    elif color == 'red':
        fColor = Fore.RED
    elif color == 'green':
        fColor = Fore.GREEN
    elif color == 'yellow':
        fColor = Fore.YELLOW
        
    print(fColor + string + Style.RESET_ALL)


def main():
    colorama.init()

    while(True):
        warehouse_to_check = select_a_warehouse()
        if warehouse_to_check == 1:
            check_warehouse(['RAW OUT'], '001')
        elif warehouse_to_check == 2:
            check_warehouse(['INGREDIENT OUT'], '002')
        elif warehouse_to_check == 3:
            check_warehouse(['BAG OUT', 'BOX OUT'], '003')
        # elif warehouse_to_check == 4:
        #     check_warehouse(['RAW OUT'], '001')
        #     check_warehouse(['INGREDIENT OUT'], '002')
        #     check_warehouse(['BAG OUT', 'BOX OUT'], '003')
        print("Press any key to continue...")
        input()


def select_a_warehouse():
    printColor('cyan', 'Select a warehouse to check:')
    printColor('green', '1 - 001 RAW MATERIALS')
    printColor('green', '2 - 002 INGREDIENTS')
    printColor('green', '3 - 003 BAGS & BOXS')
    # printColor('green', '4 - All')

    selected_index = 0
    while selected_index < 1 or selected_index > 3:
        selected_index = get_input_and_validate()
    
    return selected_index


def get_input_and_validate():
    try:
        selected_index = int(input('Select a worksheet: '))
    except:
        selected_index = 0
    
    return selected_index    


def check_warehouse(excel_tabs, u9_warehouse_num):
    items = {}
    for tab in excel_tabs:
        items = read_excel(tab, items)
    # print(items)
    items = read_u9(u9_warehouse_num, items)
    # print(items)
    items = preprocess_items(items)
    # print(items)
    show_result(items)


def read_excel(tab, items):
    filename = read_first_line('./Inventory file name.txt')
    path_of_excel = f'../{filename}.xlsx'
    wb = openpyxl.load_workbook(filename=path_of_excel, read_only=True, data_only=True)
    ws = wb[tab]
    
    for row in ws.iter_rows(min_row=1, max_row=1, min_col=1, max_col=ws.max_column):
        for cell in row:
            if cell.value == '参考料号2':
                id_col = cell.column
            elif cell.value == 'scrap':
                scrap_col = cell.column
            elif cell.value == 'inventory on hand':
                inventory_col = cell.column
                break
    
    empty_cell_count = 0
    last_row_enpty = False
    
    print(f'Reading Excel-{tab}...')
    
    for row in tqdm(range(2, ws.max_row)):
        item_id = ws.cell(row=row, column=id_col).value
        scrap_qty = ws.cell(row=row, column=scrap_col).value
        inventory = ws.cell(row=row, column=inventory_col).value
        if scrap_qty == None: scrap_qty = 0
        if inventory == None: inventory = 0
        excel_qty = round(scrap_qty + inventory, 3)
    
        if item_id == None or item_id == '' or item_id == '合计':
            if last_row_enpty: empty_cell_count += 1
            last_row_enpty = True
            if empty_cell_count >= 20:
                break
            continue
        
        if item_id == 'B35':
            item_id = 'B22'
        
        last_row_enpty = False
        
        item_id = item_id.upper()
        
        if item_id not in items:
            items[item_id] = {}
        
        if 'excel_qty' in items[item_id]:
            items[item_id]['excel_qty'] += excel_qty
        else:
            items[item_id]['excel_qty'] = excel_qty
    
    return items


def read_first_line(file_path):
    with open(file_path, 'r') as file:
        first_line = file.readline().strip()
    return first_line
        

def read_u9(u9_warehouse_num, items):
    wb = openpyxl.load_workbook(filename=f'./u9_excel/{u9_warehouse_num}.xlsx', read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]

    for row in ws.iter_rows(min_row=1, max_row=3, min_col=1, max_col=ws.max_column):
        for cell in row:
            if cell.value == '参考料号2' or cell.value == '代码':
                id_col = cell.column
                header_row = cell.row
            elif cell.value == '现存量(库存单位)':
                qty_col = cell.column
                break
    
    empty_cell_count = 0
    last_row_enpty = False
    
    if id_col == None or header_row == None or qty_col == None:
        raise 'error reading header'
    
    print(f'Reading U9-{u9_warehouse_num}...')
    
    current_row = header_row
    
    for row in tqdm(ws.iter_rows(min_row=header_row+1, max_row=1000)):
        current_row += 1
        item_id = ws.cell(current_row, column=id_col).value
        u9_qty = ws.cell(current_row, column=qty_col).value
    
        if item_id == None or item_id == '':
            if last_row_enpty: empty_cell_count += 1
            last_row_enpty = True
            if empty_cell_count >= 20:
                break
            continue
        
        if item_id == 'B35':
            item_id = 'B22'
        
        item_id = item_id.upper()
        
        last_row_enpty = False
        
        if item_id not in items:
            items[item_id] = {}
        
        if 'u9_qty' in items[item_id]:
            items[item_id]['u9_qty'] += u9_qty
        else:
            items[item_id]['u9_qty'] = u9_qty
    
    return items


def preprocess_items(items):
    new_items = copy.deepcopy(items)
    for key, value in items.items():
        if '-' not in key: continue
        match = re.search(r"^\s*([^\s-]+)", key)
        parent_key = match.group(1).strip()
        
        if 'excel_qty' in value and 'u9_qty' not in value:
            if parent_key not in new_items: continue
            new_items[parent_key]['excel_qty'] += items[key]['excel_qty']
            del new_items[key]
        
        if 'excel_qty' not in value and 'u9_qty' in value:
            if parent_key not in new_items: continue
            new_items[parent_key]['u9_qty'] += items[key]['u9_qty']
            del new_items[key]
    
    return new_items

def show_result(items):
    unmatch_items = {}
    
    for key, value in items.items():
        e = value.get("excel_qty", 0)
        u = value.get("u9_qty", 0)
        
        color = Fore.CYAN
        
        if abs(e - u) >= 1:
            color = Fore.RED
            unmatch_items[key] = value
            
        if not (e == 0 and u == 0):
            print(key + color + f' -> Excel: {e} U9: {u}' + Style.RESET_ALL)
            time.sleep(0.15)
    
    if len(unmatch_items) == 0: return
    
    output_table = []
    for key, value in unmatch_items.items():
        e = value.get("excel_qty", 0)
        u = value.get("u9_qty", 0)
        # print(key + Fore.RED + f' -> Excel: {e} U9: {u} Diff: {round(abs(e - u),3)}' + Style.RESET_ALL)
        output_table.append([key, f'Excel: {e}', f'U9: {u}', f'Diff: {round(e - u, 4)}'])
        # output_table.append([key, e, u, round(abs(e - u),3)])
        
    print('\nFound the following unmatched numbers:')
    print(tabulate(output_table))
    
# def printTabs(string):
#     length = len(str(string))

#     if length > 9:
#         return '\t'
#     if length > 1:
#         return '\t\t'
    
#     return '\t\t\t'


if __name__ == '__main__':
    main()